from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook
from uuid import uuid4
from datetime import datetime
import os
import json
import base64
import tempfile

try:
    import gspread
    from google.oauth2.service_account import Credentials
except Exception:
    # gspread/google-auth are optional and only required when Google Sheets is enabled
    gspread = None
    Credentials = None

app = FastAPI()
print("🚀 FastAPI is initializing...")


# Allow all origins for now (customize later)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = "Training Feedback.xlsx"  # Excel file in the same directory

# ----------------------------
# Google Sheets Integration
# ----------------------------
def _google_sheets_enabled() -> bool:
    """Return True if Google Sheets storage is enabled via env vars."""
    return os.getenv("GOOGLE_SHEETS_ENABLED", "false").lower() == "true" and os.getenv("GOOGLE_SPREADSHEET_ID") is not None and gspread is not None and Credentials is not None


def _load_service_account_credentials():
    """Load Google credentials from env var GOOGLE_SERVICE_ACCOUNT_JSON or JSON file.

    Supports either raw JSON, base64-encoded JSON, or direct JSON file.
    """
    # Try environment variable first
    if "GOOGLE_SERVICE_ACCOUNT_JSON" in os.environ:
        raw = os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"]
        try:
            # If it's base64 encoded, decode it
            if not raw.strip().startswith("{"):
                raw = base64.b64decode(raw).decode("utf-8")
            info = json.loads(raw)
            return Credentials.from_service_account_info(info, scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive",
            ])
        except Exception as ex:
            print(f"Warning: Could not parse GOOGLE_SERVICE_ACCOUNT_JSON env var: {ex}")
    
    # Fallback: try to read from JSON file
    try:
        json_file = "feedback-form-469519-88569fc634a1.json"
        if os.path.exists(json_file):
            with open(json_file, 'r') as f:
                info = json.load(f)
            return Credentials.from_service_account_info(info, scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive",
            ])
    except Exception as ex:
        print(f"Warning: Could not read JSON file: {ex}")
    
    raise RuntimeError("No valid Google service account credentials found")


def _get_worksheet():
    """Return the target Google Sheets worksheet instance.

    Requires env vars:
      - GOOGLE_SPREADSHEET_ID (Spreadsheet ID)
      - GOOGLE_WORKSHEET_NAME (optional, defaults to 'Sheet1')
    """
    creds = _load_service_account_credentials()
    gc = gspread.authorize(creds)
    spreadsheet_id = os.environ["GOOGLE_SPREADSHEET_ID"]
    worksheet_name = os.getenv("GOOGLE_WORKSHEET_NAME", "Sheet1")
    sh = gc.open_by_key(spreadsheet_id)
    try:
        ws = sh.worksheet(worksheet_name)
    except gspread.WorksheetNotFound:
        # Create the worksheet with headers if it doesn't exist
        ws = sh.add_worksheet(title=worksheet_name, rows="1000", cols="20")
        ws.append_row([
            "Timestamp",
            "Submission ID",
            "Full Name",
            "Email",
            "Job Role",
            "Training Title",
            "Instructor Name",
            "Content Avg",
            "Trainer Avg",
            "Organization Avg",
            "Overall Avg",
            "Covered Topics",
            "Other Topic",
            "Comments",
        ])
    return ws

# Define expected data from frontend
class FeedbackForm(BaseModel):
    full_name: str
    email: str
    job_role: str
    training_title: str
    instructor_name: str
    content_ratings: list[int]
    trainer_ratings: list[int]
    organization_ratings: list[int]
    overall_ratings: list[int]
    covered_topics: list[str]
    other_topic: str
    comments: str

@app.get("/")
async def root():
    return {"message": "Training Feedback API is running!", "endpoints": ["/submit-feedback", "/view-data", "/download-excel", "/docs"]}

@app.get("/health")
async def health_check():
    """Health check endpoint for monitoring and wake-up detection"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "service": "Training Feedback API"
    }

@app.get("/sheets-status")
async def sheets_status():
    """Check Google Sheets integration status"""
    try:
        if _google_sheets_enabled():
            ws = _get_worksheet()
            sheet_values = ws.get_all_values()
            return {
                "status": "success",
                "sheets_enabled": True,
                "spreadsheet_id": os.getenv("GOOGLE_SPREADSHEET_ID"),
                "worksheet_name": os.getenv("GOOGLE_WORKSHEET_NAME", "Sheet1"),
                "total_rows": len(sheet_values),
                "total_submissions": len([row for row in sheet_values[1:] if any(cell for cell in row)]) if len(sheet_values) > 1 else 0
            }
        else:
            return {
                "status": "success",
                "sheets_enabled": False,
                "message": "Google Sheets not configured. Check environment variables."
            }
    except Exception as e:
        return {
            "status": "error",
            "sheets_enabled": False,
            "message": f"Error checking Google Sheets: {str(e)}"
        }

@app.get("/view-data")
async def view_data():
    """View all stored feedback data from Excel or Google Sheets."""
    try:
        if _google_sheets_enabled():
            ws = _get_worksheet()
            sheet_values = ws.get_all_values()
            if not sheet_values:
                return {"status": "success", "total_submissions": 0, "headers": [], "data": []}
            headers = sheet_values[0]
            data = [row for row in sheet_values[1:] if any(cell for cell in row)]
            return {
                "status": "success",
                "total_submissions": len(data),
                "headers": headers,
                "data": data,
            }

        # Fallback: Excel file storage
        if not os.path.exists(EXCEL_FILE):
            return {"status": "error", "message": f"{EXCEL_FILE} not found."}

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        data = []
        headers = []

        # Get headers (first row)
        for cell in ws[1]:
            headers.append(cell.value)

        # Get all data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):  # Skip empty rows
                data.append(row)

        return {
            "status": "success",
            "total_submissions": len(data),
            "headers": headers,
            "data": data,
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/download-excel")
async def download_excel():
    """Download the Excel file with all stored data from Google Sheets or local Excel"""
    try:
        if _google_sheets_enabled():
            # Create Excel file from Google Sheets data
            ws = _get_worksheet()
            sheet_values = ws.get_all_values()
            
            if not sheet_values:
                return {"status": "error", "message": "No data found in Google Sheets."}
            
            # Create Excel file in memory
            wb = Workbook()
            ws_excel = wb.active
            
            # Add headers
            headers = sheet_values[0]
            ws_excel.append(headers)
            
            # Add data rows
            for row in sheet_values[1:]:
                if any(cell for cell in row):  # Skip empty rows
                    ws_excel.append(row)
            
            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            temp_file.close()
            
            return FileResponse(
                path=temp_file.name,
                filename="Training_Feedback_Data_Google_Sheets.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # Fallback to local Excel file
        if not os.path.exists(EXCEL_FILE):
            return {"status": "error", "message": f"{EXCEL_FILE} not found."}
        
        return FileResponse(
            path=EXCEL_FILE,
            filename="Training_Feedback_Data.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.delete("/delete-feedback/{submission_id}")
async def delete_feedback(submission_id: str):
    """Delete a specific feedback entry by submission ID."""
    try:
        if _google_sheets_enabled():
            ws = _get_worksheet()
            # Find submission_id in column B (2)
            cell = ws.find(submission_id)
            if not cell:
                return {"status": "error", "message": f"Submission with ID {submission_id} not found."}
            ws.delete_rows(cell.row)
            return {"status": "success", "message": f"Submission {submission_id} deleted successfully."}

        # Fallback: Excel file
        if not os.path.exists(EXCEL_FILE):
            return {"status": "error", "message": f"{EXCEL_FILE} not found."}

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Find the row with the matching submission ID
        row_to_delete = None
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[1] == submission_id:  # submission_id is in column B (index 1)
                row_to_delete = row_num
                break

        if row_to_delete is None:
            return {"status": "error", "message": f"Submission with ID {submission_id} not found."}

        # Delete the row
        ws.delete_rows(row_to_delete)
        wb.save(EXCEL_FILE)

        return {"status": "success", "message": f"Submission {submission_id} deleted successfully."}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/submit-feedback")
async def submit_feedback(form: FeedbackForm):
    try:
        submission_id = str(uuid4())[:8]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        def average(lst):
            return round(sum(lst) / len(lst), 2) if lst else 0

        content_avg = average(form.content_ratings)
        trainer_avg = average(form.trainer_ratings)
        org_avg = average(form.organization_ratings)
        overall_avg = average(form.overall_ratings)

        row = [
            timestamp,
            submission_id,
            form.full_name,
            form.email,
            form.job_role,
            form.training_title,
            form.instructor_name,
            content_avg,
            trainer_avg,
            org_avg,
            overall_avg,
            ", ".join(form.covered_topics),
            form.other_topic,
            form.comments,
        ]

        if _google_sheets_enabled():
            ws = _get_worksheet()
            ws.append_row(row)
            return {"status": "success", "submission_id": submission_id, "storage": "sheets"}

        # Fallback to Excel storage
        if not os.path.exists(EXCEL_FILE):
            # If Excel file is missing, initialize with headers
            wb = Workbook()
            ws = wb.active
            ws.append([
                "Timestamp",
                "Submission ID",
                "Full Name",
                "Email",
                "Job Role",
                "Training Title",
                "Instructor Name",
                "Content Avg",
                "Trainer Avg",
                "Organization Avg",
                "Overall Avg",
                "Covered Topics",
                "Other Topic",
                "Comments",
            ])
            wb.save(EXCEL_FILE)

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(row)
        wb.save(EXCEL_FILE)

        return {"status": "success", "submission_id": submission_id, "storage": "excel"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=9000)

print("✅ FastAPI loaded successfully")
